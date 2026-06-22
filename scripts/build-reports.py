#!/usr/bin/env python3
"""日々の活動報告.xlsx と 活動報告　写真 から reports.json を生成します。"""
from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "日々の活動報告.xlsx"
PHOTO_SRC = ROOT / "活動報告　写真"
WEBSITE = ROOT / "website"
PHOTO_DST = WEBSITE / "assets" / "reports"
JSON_OUT = WEBSITE / "data" / "reports.json"

RETENTION_MONTHS = 3
WEEKDAYS_JA = ("月", "火", "水", "木", "金", "土", "日")

HEADER_ALIASES = {
    "date": {"日付", "date", "Date"},
    "title": {"タイトル", "題名", "title", "Title"},
    "author": {"担当", "担当者", "著者", "活動報告担当", "author", "Author"},
    "comment": {"コメント", "comment", "Comment", "一言", "感想"},
    "body": {"本文", "内容", "活動内容", "活動報告", "レポート", "body", "Body"},
    "photo": {"写真", "photo", "Photo", "画像"},
}


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value) -> date | None:
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date):
                return parsed
        except (ValueError, OverflowError, OSError):
            pass

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    match = re.search(r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))

    return None


def resolve_report_date(date_value, photo_ref: str | None) -> date | None:
    if photo_ref:
        photo_date = photo_date_from_name(photo_ref)
        if photo_date:
            return photo_date
    return parse_date(date_value)


def date_label(value: date) -> str:
    return f"{value.year}年{value.month}月{value.day}日（{WEEKDAYS_JA[value.weekday()]}）"


def parse_photo_ref(value, fallback_date: date | None = None) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().replace("\u3000", " ")
    if re.fullmatch(r"\d{8}", text):
        return f"photo {text}.jpg"
    if re.fullmatch(r"\d{4}", text):
        return f"photo {text}.jpg"
    if not text.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
        text = f"{text}.jpg"
    return text


def photo_date_from_name(filename: str) -> date | None:
    stem = Path(filename).stem.replace("\u3000", " ").strip()

    match = re.search(r"(\d{8})$", stem)
    if match:
        digits = match.group(1)
        return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))

    match = re.search(r"(\d{6})", stem)
    if match:
        digits = match.group(1)
        year = 2000 + int(digits[:2])
        month = int(digits[2:4])
        day = int(digits[4:6])
        try:
            return date(year, month, day)
        except ValueError:
            pass

    match = re.search(r"(\d{4})$", stem)
    if match:
        digits = match.group(1)
        month = int(digits[:2])
        day = int(digits[2:])
        year = date.today().year
        try:
            return date(year, month, day)
        except ValueError:
            return None

    return None


def find_header_map(row: list) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        text = cell_text(value)
        for key, aliases in HEADER_ALIASES.items():
            if text in aliases:
                mapping[key] = index
    if "date" in mapping and ("comment" in mapping or "body" in mapping or "title" in mapping):
        return mapping
    return None


def read_reports_from_excel() -> dict[date, dict]:
    if not XLSX.exists():
        return {}

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    reports: dict[date, dict] = {}
    header_map: dict[str, int] | None = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not any(row):
            continue

        if header_map is None:
            header_map = find_header_map(list(row))
            if header_map:
                continue

        if header_map is None:
            photo_ref = parse_photo_ref(row[3] if len(row) > 3 else row[4] if len(row) > 4 else None)
            report_date = resolve_report_date(row[0] if len(row) > 0 else None, photo_ref)
            if report_date is None:
                continue
            has_title_column = len(row) > 4 and parse_photo_ref(row[4]) is not None
            reports[report_date] = {
                "date": report_date,
                "title": cell_text(row[1]) if has_title_column else "",
                "author": cell_text(row[2] if has_title_column else row[1]),
                "comment": cell_text(row[3] if has_title_column else row[2]),
                "body": "",
                "photo_ref": photo_ref,
            }
            continue

        photo_ref = parse_photo_ref(
            row[header_map["photo"]] if "photo" in header_map and header_map["photo"] < len(row) else None
        )
        report_date = resolve_report_date(
            row[header_map["date"]] if header_map["date"] < len(row) else None,
            photo_ref,
        )
        if report_date is None:
            continue

        comment = ""
        if "comment" in header_map and header_map["comment"] < len(row):
            comment = cell_text(row[header_map["comment"]])
        body = ""
        if "body" in header_map and header_map["body"] < len(row):
            body = cell_text(row[header_map["body"]])

        reports[report_date] = {
            "date": report_date,
            "title": cell_text(row[header_map["title"]]) if "title" in header_map else "",
            "author": cell_text(row[header_map["author"]]) if "author" in header_map else "",
            "comment": comment or body,
            "body": body,
            "photo_ref": photo_ref,
        }

    return reports


def scan_source_photos() -> dict[date, Path]:
    photos: dict[date, Path] = {}
    if not PHOTO_SRC.exists():
        return photos

    for path in PHOTO_SRC.iterdir():
        if path.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
            continue
        report_date = photo_date_from_name(path.name)
        if report_date is None:
            print(f"警告: 日付を読み取れない写真をスキップしました: {path.name}")
            continue
        photos[report_date] = path

    return photos


def default_title(report_date: date) -> str:
    return f"{report_date.month}月{report_date.day}日 金曜アクト"


def default_comment(report_date: date) -> str:
    return f"{date_label(report_date)}のアクトの活動報告です。"


def format_author(author: str) -> str:
    text = author.strip()
    if not text:
        return "活動報告担当"
    if text.startswith("活動報告担当"):
        return text
    return f"活動報告担当：{text}"


def retention_cutoff(today: date | None = None) -> date:
    today = today or date.today()
    return today - timedelta(days=RETENTION_MONTHS * 30)


def copy_photo(src: Path, report_date: date) -> str:
    PHOTO_DST.mkdir(parents=True, exist_ok=True)
    ext = src.suffix.lower()
    if ext == ".jpeg":
        ext = ".jpg"
    dst_name = f"report-{report_date.isoformat()}{ext}"
    dst_path = PHOTO_DST / dst_name
    shutil.copy2(src, dst_path)
    return f"assets/reports/{dst_name}"


def cleanup_destination(keep_paths: set[Path]) -> None:
    if not PHOTO_DST.exists():
        return
    for path in PHOTO_DST.iterdir():
        if path.is_file() and path.resolve() not in keep_paths:
            path.unlink()


def build_reports() -> list[dict]:
    excel_reports = read_reports_from_excel()
    photo_reports = scan_source_photos()
    cutoff = retention_cutoff()
    used_dest_paths: set[Path] = set()
    merged_dates = sorted(set(excel_reports) | set(photo_reports), reverse=True)
    reports: list[dict] = []

    for report_date in merged_dates:
        if report_date < cutoff:
            continue

        excel_entry = excel_reports.get(report_date, {})
        photo_path = photo_reports.get(report_date)
        photo_ref = excel_entry.get("photo_ref")

        if photo_path is None and photo_ref:
            candidate = PHOTO_SRC / photo_ref
            if candidate.exists():
                photo_path = candidate
            else:
                for path in PHOTO_SRC.iterdir():
                    if path.name.lower() == photo_ref.lower():
                        photo_path = path
                        break

        if photo_path is None:
            continue

        title = excel_entry.get("title") or default_title(report_date)
        author = format_author(excel_entry.get("author") or "")
        comment = (
            excel_entry.get("comment")
            or excel_entry.get("body")
            or default_comment(report_date)
        )
        photo = copy_photo(photo_path, report_date)
        used_dest_paths.add((PHOTO_DST / Path(photo).name).resolve())

        reports.append(
            {
                "id": report_date.isoformat(),
                "date": report_date.isoformat(),
                "dateLabel": date_label(report_date),
                "title": title,
                "author": author,
                "comment": comment,
                "photo": photo,
            }
        )

    cleanup_destination(used_dest_paths)
    return reports


def main() -> None:
    reports = build_reports()
    payload = {
        "reportFrequency": "毎週金曜日のアクト",
        "siteUpdateInterval": "ホームページ更新：約3週間に1回",
        "retentionMonths": RETENTION_MONTHS,
        "reports": reports,
    }

    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"生成完了: {JSON_OUT} ({len(reports)} 件)")
    if reports:
        print(f"  掲載期間: {retention_cutoff().isoformat()} 以降")
    else:
        print("警告: 活動報告データがありません。Excel または 活動報告　写真 を確認してください。")


if __name__ == "__main__":
    main()
